"""
APS智能排产演示系统 - FastAPI 后端
"""
import json
import os
import csv
import io
import datetime
from typing import Optional

from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from config import APP_TITLE, APP_VERSION, STATIC_DIR
from database import init_db, query, execute
from optimizer import SimplifiedALNS


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title=APP_TITLE, version=APP_VERSION, lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==================== 数据模型 ====================

class OrderCreate(BaseModel):
    product_id: int
    quantity: int
    deadline: str = ""
    priority: int = 5
    customer: str = ""
    remark: str = ""


class OrderPriority(BaseModel):
    priority: int


class ConstraintUpdate(BaseModel):
    constraint_name: str
    min_batch: Optional[int] = None
    max_batch: Optional[int] = None
    gap_limit: Optional[int] = None
    weight: Optional[float] = None
    is_active: Optional[int] = None


# ==================== 排产计划 API（核心） ====================

@app.post("/api/schedule/optimize")
async def optimize_schedule(data: Optional[dict] = None):
    """
    执行排产优化
    - data.order_ids: 可选，指定排产的订单ID列表（逗号分隔），不传则排所有待排产订单
    """
    ids = None
    if data and "order_ids" in data and data["order_ids"]:
        ids = [int(x.strip()) for x in data["order_ids"].split(",") if x.strip()]

    optimizer = SimplifiedALNS(order_ids=ids)
    result = optimizer.optimize()

    if result is None:
        raise HTTPException(status_code=400, detail="没有待排产的订单数据")

    # 保存排产结果
    result_id = execute(
        "INSERT INTO schedule_result (order_sequence, score, "
        "switch_count_before, switch_count_after, "
        "batch_rate_before, batch_rate_after, "
        "gap_rate_before, gap_rate_after) VALUES (?,?,?,?,?,?,?,?)",
        [
            json.dumps(result["order_sequence"]),
            result["after"]["score"],
            result["before"]["switch_count"],
            result["after"]["switch_count"],
            result["before"]["batch_rate"],
            result["after"]["batch_rate"],
            result["before"]["gap_rate"],
            result["after"]["gap_rate"],
        ],
    )

    # 保存排产明细
    for item in result["sequence_detail"]:
        execute(
            "INSERT INTO schedule_detail (result_id, sequence_no, order_id, "
            "product_name, attr_a, attr_b, composite_craft, special_component, "
            "quantity, batch_id) VALUES (?,?,?,?,?,?,?,?,?,?)",
            [
                result_id,
                item["order_id"],
                item["order_id"],
                item["product_name"],
                item["attr_a"],
                item["attr_b"],
                item["composite_craft"],
                item["special_component"],
                item["quantity"],
                item["batch_id"],
            ],
        )

    # 更新订单状态
    for oid in result["order_sequence"]:
        execute('UPDATE "order" SET status = 1 WHERE id = ?', [oid])

    return {"result_id": result_id, "data": result, "message": "排产优化完成"}


@app.get("/api/schedule/result/{result_id}")
async def get_schedule_result(result_id: int):
    row = query("SELECT * FROM schedule_result WHERE id = ?", [result_id], one=True)
    if not row:
        raise HTTPException(status_code=404, detail="排产结果不存在")
    details = query(
        "SELECT * FROM schedule_detail WHERE result_id = ? ORDER BY sequence_no",
        [result_id],
    )
    row["details"] = details
    row["order_sequence"] = json.loads(row["order_sequence"])
    return row


@app.get("/api/schedule/results")
async def list_schedule_results():
    rows = query("SELECT id, score, switch_count_before, switch_count_after, "
                 "batch_rate_before, batch_rate_after, create_time "
                 "FROM schedule_result ORDER BY create_time DESC")
    return rows


@app.get("/api/schedule/export/{result_id}")
async def export_schedule(result_id: int):
    """导出排产结果为CSV"""
    row = query("SELECT * FROM schedule_result WHERE id = ?", [result_id], one=True)
    if not row:
        raise HTTPException(status_code=404, detail="结果不存在")

    details = query(
        "SELECT sd.*, o.order_no, o.deadline, o.priority "
        "FROM schedule_detail sd "
        "JOIN \"order\" o ON sd.order_id = o.id "
        "WHERE sd.result_id = ? ORDER BY sd.sequence_no",
        [result_id],
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["顺序号", "订单号", "产品名称", "属性A", "属性B",
                     "复合工艺", "特殊组件", "数量", "交期", "优先级", "批次号"])
    for i, d in enumerate(details, 1):
        writer.writerow([
            i, d["order_no"], d["product_name"], d["attr_a"], d["attr_b"],
            d["composite_craft"], d["special_component"], d["quantity"],
            d["deadline"], d["priority"], d["batch_id"],
        ])

    output.seek(0)
    content = '﻿' + output.getvalue()  # BOM for Excel
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=schedule_result_{result_id}.csv"},
    )


@app.get("/api/schedule/gantt-data/{result_id}")
async def get_gantt_data(result_id: int):
    """获取甘特图展示数据"""
    row = query("SELECT * FROM schedule_result WHERE id = ?", [result_id], one=True)
    if not row:
        raise HTTPException(status_code=404, detail="结果不存在")

    details = query(
        "SELECT sd.*, o.order_no, o.deadline "
        "FROM schedule_detail sd "
        "JOIN \"order\" o ON sd.order_id = o.id "
        "WHERE sd.result_id = ? ORDER BY sd.sequence_no",
        [result_id],
    )

    gantt_data = []
    current_time = datetime.datetime.now()
    for i, d in enumerate(details):
        start = current_time + datetime.timedelta(hours=i * 2)
        end = start + datetime.timedelta(hours=1)
        gantt_data.append({
            "task_id": i + 1,
            "order_no": d["order_no"],
            "product_name": d["product_name"],
            "attr_a": d["attr_a"],
            "attr_b": d["attr_b"],
            "composite_craft": d["composite_craft"],
            "special_component": d["special_component"],
            "start": start.strftime("%Y-%m-%d %H:%M"),
            "end": end.strftime("%Y-%m-%d %H:%M"),
            "batch_id": d["batch_id"],
            "quantity": d["quantity"],
            "priority": d.get("priority", 5),
        })
    return gantt_data


# ==================== 订单 API ====================

@app.get("/api/orders")
async def list_orders(status: Optional[int] = None, page: int = 1, page_size: int = 50):
    where = ""
    params = []
    if status is not None:
        where = "WHERE o.status = ?"
        params.append(status)
    total = query(f'SELECT COUNT(*) as cnt FROM "order" o {where}', params, one=True)
    offset = (page - 1) * page_size
    rows = query(
        f'SELECT o.*, p.name as product_name FROM "order" o '
        f'JOIN product p ON o.product_id = p.id {where} '
        f"ORDER BY o.id LIMIT ? OFFSET ?",
        params + [page_size, offset],
    )
    return {"total": total["cnt"], "page": page, "page_size": page_size, "data": rows}


@app.get("/api/orders/{order_id}")
async def get_order(order_id: int):
    row = query(
        f'SELECT o.*, p.name as product_name, p.attr_a, p.attr_b, '
        f'p.composite_craft, p.special_component, p.model_type '
        f'FROM "order" o JOIN product p ON o.product_id = p.id '
        f"WHERE o.id = ?",
        [order_id],
        one=True,
    )
    if not row:
        raise HTTPException(status_code=404, detail="订单不存在")
    return row


@app.post("/api/orders")
async def create_order(order: OrderCreate):
    order_no = f"ORD-{datetime.date.today().strftime('%Y%m%d')}-{datetime.datetime.now().microsecond:04d}"
    oid = execute(
        'INSERT INTO "order" (order_no, product_id, quantity, deadline, priority, customer, remark) '
        "VALUES (?,?,?,?,?,?,?)",
        [order_no, order.product_id, order.quantity, order.deadline,
         order.priority, order.customer, order.remark],
    )
    return {"id": oid, "order_no": order_no, "message": "订单创建成功"}


@app.put("/api/orders/{order_id}/priority")
async def set_order_priority(order_id: int, data: OrderPriority):
    execute('UPDATE "order" SET priority = ? WHERE id = ?', [data.priority, order_id])
    return {"message": "优先级更新成功"}


@app.post("/api/orders/import")
async def import_orders(file: UploadFile = File(...)):
    """导入订单（CSV/Excel）"""
    content = await file.read()
    count = 0

    if file.filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            product_name = row.get("产品名称", "").strip()
            product = query("SELECT id FROM product WHERE name = ?", [product_name], one=True)
            if not product:
                continue
            # 使用时间戳确保订单号唯一性
            timestamp = datetime.datetime.now().strftime("%H%M%S%f")[:6]  # 取微秒前6位
            order_no = f"IMP-{datetime.date.today().strftime('%Y%m%d')}-{timestamp}-{count:02d}"
            execute(
                'INSERT INTO "order" (order_no, product_id, quantity, deadline, priority, customer) '
                "VALUES (?,?,?,?,?,?)",
                [order_no, product["id"],
                 int(row.get("数量", 0)),
                 row.get("交期", ""),
                 int(row.get("优先级", 5)),
                 row.get("客户", "")],
            )
            count += 1

    elif file.filename.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        if not ws:
            return {"count": 0, "message": "Excel文件为空"}

        # Read header row to find column indices
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        col_map = {h: i for i, h in enumerate(headers)}

        for r in range(2, ws.max_row + 1):
            product_name = str(ws.cell(r, col_map.get("产品名称", 1) + 1).value or "").strip()
            if not product_name:
                continue
            product = query("SELECT id FROM product WHERE name = ?", [product_name], one=True)
            if not product:
                continue
            qty = int(ws.cell(r, col_map.get("数量", 8) + 1).value or 0)
            deadline = str(ws.cell(r, col_map.get("交期", 9) + 1).value or "")
            priority = int(ws.cell(r, col_map.get("优先级", 10) + 1).value or 5)
            customer = str(ws.cell(r, col_map.get("客户", 11) + 1).value or "")
            # 使用时间戳确保订单号唯一性
            timestamp = datetime.datetime.now().strftime("%H%M%S%f")[:6]  # 取微秒前6位
            order_no = f"XLS-{datetime.date.today().strftime('%Y%m%d')}-{timestamp}-{count:02d}"
            execute(
                'INSERT INTO "order" (order_no, product_id, quantity, deadline, priority, customer) '
                "VALUES (?,?,?,?,?,?)",
                [order_no, product["id"], qty, deadline, priority, customer],
            )
            count += 1
        wb.close()

    else:
        return {"message": "请上传CSV或Excel文件", "count": 0}

    return {"count": count, "message": f"成功导入 {count} 条订单"}


# ==================== 约束配置 API ====================

@app.get("/api/constraints")
async def list_constraints():
    return query("SELECT * FROM constraint_config ORDER BY id")


@app.put("/api/constraints")
async def update_constraints(constraints: list[ConstraintUpdate]):
    for c in constraints:
        sets = []
        params = []
        if c.min_batch is not None:
            sets.append("min_batch = ?")
            params.append(c.min_batch)
        if c.max_batch is not None:
            sets.append("max_batch = ?")
            params.append(c.max_batch)
        if c.gap_limit is not None:
            sets.append("gap_limit = ?")
            params.append(c.gap_limit)
        if c.weight is not None:
            sets.append("weight = ?")
            params.append(c.weight)
        if c.is_active is not None:
            sets.append("is_active = ?")
            params.append(c.is_active)
        if sets:
            params.append(c.constraint_name)
            execute(f"UPDATE constraint_config SET {', '.join(sets)}, "
                    f"update_time = CURRENT_TIMESTAMP WHERE constraint_name = ?", params)
    return {"message": "约束配置已更新"}


# ==================== 产品 API ====================

@app.get("/api/products")
async def list_products():
    return query("SELECT * FROM product ORDER BY id")


# ==================== 假数据模拟 API ====================

def make_mock_data(table):
    return query(f"SELECT * FROM {table} ORDER BY id")


@app.get("/api/dashboard/employees")
async def get_employees():
    return make_mock_data("employee")


@app.get("/api/dashboard/equipment")
async def get_equipment():
    return make_mock_data("equipment")


@app.get("/api/dashboard/materials")
async def get_materials():
    return make_mock_data("material")


@app.get("/api/dashboard/mes-lines")
async def get_mes_lines():
    return [
        {"id": 1, "line_code": "L001", "line_name": "产线一", "status": 1,
         "product_name": "设备一", "today_qty": 45, "plan_qty": 120, "speed": 87.5},
        {"id": 2, "line_code": "L002", "line_name": "产线二", "status": 1,
         "product_name": "设备二", "today_qty": 38, "plan_qty": 100, "speed": 92.3},
        {"id": 3, "line_code": "L003", "line_name": "产线三", "status": 0,
         "product_name": "设备三", "today_qty": 0, "plan_qty": 80, "speed": 0},
        {"id": 4, "line_code": "L004", "line_name": "产线四", "status": 2,
         "product_name": "设备四", "today_qty": 12, "plan_qty": 90, "speed": 45.0},
    ]


@app.get("/api/dashboard/quality")
async def get_quality_data():
    """质量监控数据"""
    import random
    random.seed(42)
    dates = [(datetime.date.today() - datetime.timedelta(days=i)).strftime("%m-%d")
             for i in range(14)]
    return {
        "dates": dates[::-1],
        "yield_rate": [round(random.uniform(92, 99.5), 1) for _ in range(14)],
        "defect_count": [random.randint(1, 15) for _ in range(14)],
    }


@app.get("/api/dashboard/safety")
async def get_safety_data():
    return make_mock_data("employee")


@app.get("/api/dashboard/anomaly")
async def get_anomaly_data():
    return [
        {"id": 1, "type": "温度异常", "source": "设备三", "time": "2026-05-16 08:23", "level": "警告", "status": 0},
        {"id": 2, "type": "振动超标", "source": "设备一", "time": "2026-05-16 07:15", "level": "严重", "status": 1},
        {"id": 3, "type": "物料短缺", "source": "物料B", "time": "2026-05-15 22:00", "level": "警告", "status": 1},
        {"id": 4, "type": "良率下降", "source": "产线二", "time": "2026-05-15 14:30", "level": "提示", "status": 0},
    ]


@app.get("/api/dashboard/mock-chart")
async def get_mock_chart():
    """通用的图表mock数据"""
    import random
    random.seed(123)
    return {
        "order_completion": {
            "rate": 87.5,
            "trend": [random.uniform(75, 95) for _ in range(12)],
            "labels": ["1月", "2月", "3月", "4月", "5月", "6月",
                       "7月", "8月", "9月", "10月", "11月", "12月"],
        },
        "oee": [
            {"name": "设备一", "oee": 85.5, "availability": 92.0, "performance": 95.3, "quality": 97.8},
            {"name": "设备二", "oee": 92.1, "availability": 96.5, "performance": 97.0, "quality": 98.2},
            {"name": "设备三", "oee": 0, "availability": 0, "performance": 0, "quality": 0},
            {"name": "设备四", "oee": 78.9, "availability": 85.0, "performance": 93.5, "quality": 99.1},
            {"name": "设备五", "oee": 88.3, "availability": 93.0, "performance": 96.0, "quality": 98.5},
            {"name": "设备六", "oee": 76.2, "availability": 82.0, "performance": 94.0, "quality": 97.0},
        ],
        "production_efficiency": {
            "labels": ["第1周", "第2周", "第3周", "第4周"],
            "actual": [850, 920, 880, 950],
            "planned": [1000, 1000, 1000, 1000],
        },
        "downtime_analysis": {
            "labels": ["设备故障", "换线调整", "物料等待", "质量异常", "其他"],
            "values": [120, 85, 60, 35, 20],
        },
    }


@app.get("/api/dashboard/schedule-stats")
async def get_schedule_stats():
    latest = query("SELECT * FROM schedule_result ORDER BY id DESC LIMIT 1", one=True)
    if latest:
        before = latest["switch_count_before"]
        after = latest["switch_count_after"]
        improvement = round((before - after) / max(before, 1) * 100, 1)
        return {
            "has_data": True,
            "improvement": improvement,
            "batch_rate_before": latest["batch_rate_before"],
            "batch_rate_after": latest["batch_rate_after"],
            "switch_before": before,
            "switch_after": after,
            "score": latest["score"],
        }
    return {"has_data": False}


# ==================== 静态文件服务 ====================

if os.path.exists(STATIC_DIR):
    app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="frontend")


# ==================== 启动入口 ====================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
